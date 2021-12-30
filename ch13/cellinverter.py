import openpyxl,os

file = input("enter the absolute path of the file: ")

wb = openpyxl.load_workbook(file)
sheet = wb.active
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

rows = []

for row in range(1, sheet.max_row + 1):
    val = []
    for cell in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=cell).value
        val.append(cell_value)
    rows.append(val)

column_num = 1
for row in rows:
    row_num = 1
    for cell in row:
        sheet2.cell(row=row_num, column=column_num).value = cell
        row_num += 1
    column_num += 1

wb2.save('cellinverter.xlsx')