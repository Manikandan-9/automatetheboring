import sys,openpyxl
from openpyxl.utils import get_column_letter


n = int(sys.argv[1])


wb = openpyxl.Workbook()
sheet = wb.active

while n>0:
    sheet.cell(row=n+1,column=1).value = n
    sheet.cell(row=1,column=n+1).value = n
    n-=1

n = int(sys.argv[1])

for col in range(2,n+2):
    for row in range(2,n+2):
        col_letter = get_column_letter(col)
        sheet[col_letter + str(row)] = (col-1)*(row-1)

wb.save('multiplicationtable.xlsx')