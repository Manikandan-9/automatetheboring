import openpyxl

excelfile = input("Enter file name: ")
wb = openpyxl.load_workbook(excelfile)
sheet = wb.active

col_no=1

for row in range(1,sheet.max_row+1):
    data = []
    for col in range(1,sheet.max_column+1):
        if sheet.cell(row=row,column=col).value!=None:
            data.append(sheet.cell(row=row,column=col).value)

    file = open(str(col_no)+excelfile+'.txt','w')
    for val in data:
        file.write('{}\n'.format(val))
    file.close()
    col_no+=1